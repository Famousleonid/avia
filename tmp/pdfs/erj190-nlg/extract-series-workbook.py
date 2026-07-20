import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

source = Path(r"C:\airplane\ERJ 190 NLG\190-70745-series.xlsx")
target = Path(__file__).with_name("series-sheet-values.json")
sheets = [
    "SB", "LLP", "NDT", "NDT Feedback", "CAD", "CAD AIRCO PLATING",
    "PAINT", "PAINT (2)", "PRL",
]
pn_pattern = re.compile(r"(?<![A-Z0-9-])(?:[A-Z]{1,6}-?)?\d{2,6}(?:-[A-Z0-9]{1,8}){1,5}(?![A-Z0-9-])", re.I)

def scalar(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value

book = load_workbook(source, read_only=True, data_only=False)
output = {"source": str(source), "sheets": {}}

for name in sheets:
    sheet = book[name]
    cells = []
    candidates = []
    for row in sheet.iter_rows():
        for cell in row:
            value = scalar(cell.value)
            if value is None or value == "":
                continue
            text = str(value)
            cells.append({"cell": cell.coordinate, "value": text})
            for match in pn_pattern.finditer(text.replace("\n", " ")):
                candidates.append({"cell": cell.coordinate, "value": match.group(0).upper(), "context": text})
    output["sheets"][name] = {"cells": cells, "candidates": candidates}

target.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
print(target)
